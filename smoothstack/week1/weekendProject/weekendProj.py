from openpyxl import load_workbook
import os
import datetime as dt
import logging 

os.chdir(os.path.dirname(__file__))
logging.basicConfig(filename ='logFile.log',format="%(asctime)s:%(levelname)s:%(message)s",level= logging.INFO)


filename = "expedia_report_monthly_march_2018.xlsx"
lfn = filename.split("_")

try:
    wb = load_workbook(filename)
except:
    logging.exception("Filename does not Exist")
else:
    logging.info("File was opened sucessfully")
    ws = wb['Summary Rolling MoM']
    ws2 = wb['VOC Rolling MoM']

def getDate(l):
    d = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
        "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
    m = l[3][:3]
    y = int(l[4][:4])
    newM = d[m]
    ans = dt.datetime(y,newM,1,0,0)
    logging.info("The requested date is %s"%ans)
    return ans


date = getDate(lfn)

def summaryTab(ws,date):
    get_percentage = lambda x : x * 100
    logging.info("Summary Rolling")
    for row in ws.iter_rows(min_row=2, max_col=6, max_row=13, values_only=True):
        if row[0].year == date.year and row[0].month == date.month:
            logging.info("Calls Offered: %s"%row[1])
            logging.info("Abandon after 30s : %s"%(get_percentage(row[2]))+"%")
            logging.info("FCR : %s"%(get_percentage(row[3]))+"%")
            logging.info("DSAT : %s"%(get_percentage(row[4]))+"%")
            logging.info("CSAT : %s"%(get_percentage(row[5]))+"%")
            

summaryTab(ws,date)

def vocTab(ws,date):
    Promoter_Score = lambda x : ("Bad" if x < 200 else "Good")
    Pass_Dec_Score = lambda x : ("Bad" if x < 100 else "Good")
    get_percentage = lambda x : x * 100
    logging.info("VOC Rolling")
    for col in ws.iter_cols(min_row=1, min_col=2, max_row=19,max_col=13,values_only=True):
        if col[0].year == date.year and col[0].month == date.month:
            logging.info("Base Size : %s"%col[2])
            
            logging.info("Promoters: %s"%col[3])
            logging.info("Promoters score : %s"%(Promoter_Score(col[3])))

            logging.info("Passives: %s"%col[5])
            logging.info("Passives score : %s"%(Pass_Dec_Score(col[5])))
            
            logging.info("Decractors: %s"%col[7])
            logging.info("Decractors score : %s"%(Pass_Dec_Score(col[7])))
            
            logging.info("Overall NPS AARP Total : %s"%(get_percentage(col[12]))+"%")
            logging.info("Sat with Agent AARP Total : %s"%(get_percentage(col[15]))+"%")
            logging.info("DSat with Agent AARP Total : %s"%(get_percentage(col[18]))+"%")
            
            

vocTab(ws2,date)